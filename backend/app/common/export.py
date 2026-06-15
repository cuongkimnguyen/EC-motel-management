"""Shared helpers for streaming Excel and PDF responses.

Usage:
    wb = build_excel_workbook("Sheet1", headers, rows)
    return excel_response(wb, "expenses_2025-06.xlsx")

    html = "<html>...</html>"
    return pdf_response(html, "report_2025-06.pdf")
"""
import io
from typing import Any

import openpyxl
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_excel_workbook(
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    col_widths: list[int] | None = None,
) -> openpyxl.Workbook:
    """Build an openpyxl Workbook with a header row and data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        width = col_widths[col_idx - 1] if col_widths and col_idx <= len(col_widths) else 18
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    data_align = Alignment(vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align

    return wb


def excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    """Serialize workbook to bytes and return as a streaming download."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def pdf_response(html: str, filename: str) -> StreamingResponse:
    """Return HTML content as a downloadable PDF-like HTML file.

    NOTE: Full PDF rendering (via WeasyPrint) deferred due to system dependency
    requirements. This returns the HTML directly with PDF content-type headers
    for now. Clients can render/print the HTML as PDF.
    """
    buffer = io.BytesIO(html.encode("utf-8"))
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
